import os
import glob
import json
import cv2
import openpyxl
from openpyxl.styles import Alignment

from ocr.visualizer import detect_color_presence_bgr

def generate_excel_report(output_root="output", excel_filename="traffic_light_report_new.xlsx"):
    print(f"Generating report from {output_root}...")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Traffic Light Report"
    
    # Headers
    headers = ["Text", "Detected Colors", "Result Code", "Source Images"]
    ws.append(headers)
    
    # Column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 40
    ws.column_dimensions['D'].width = 80
    
    # Find all micro images via JSONs
    json_pattern = os.path.join(output_root, "**", "micro_img", "micro_*.json")
    json_files = glob.glob(json_pattern, recursive=True)
    
    print(f"Found {len(json_files)} JSON files.")
    
    # Aggregation Dictionary
    # Key: text
    # Value: { 'colors': set(), 'img_paths': [] }
    aggregated_data = {}
    
    for json_file in sorted(json_files):
        try:
            with open(json_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            img_filename = data.get("micro_image_name")
            text = data.get("text", "").strip()
            
            if not text:
                continue

            # Construct image path
            dir_path = os.path.dirname(json_file)
            img_path = os.path.join(dir_path, img_filename)
            
            if not os.path.exists(img_path):
                continue
                
            # Re-detect colors
            img_bgr = cv2.imread(img_path)
            if img_bgr is None:
                continue
                
            color_info = detect_color_presence_bgr(img_bgr)
            presence = color_info["presence"]
            
            if text not in aggregated_data:
                aggregated_data[text] = {
                    'colors': set(),
                    'img_paths': []
                }
            
            aggregated_data[text]['img_paths'].append(img_path)
            
            # Add detected colors to the set
            if presence.get("white"): aggregated_data[text]['colors'].add("White")
            if presence.get("red"): aggregated_data[text]['colors'].add("Red")
            if presence.get("green"): aggregated_data[text]['colors'].add("Green")
            if presence.get("yellow"): aggregated_data[text]['colors'].add("Yellow")
            if presence.get("blue"): aggregated_data[text]['colors'].add("Blue")
            
        except Exception as e:
            print(f"Error processing {json_file}: {e}")
    
    # Process aggregated data and write to Excel
    current_row = 2
    
    # Sort by text for cleaner output
    for text in sorted(aggregated_data.keys()):
        data = aggregated_data[text]
        colors_set = data['colors']
        img_paths = data['img_paths']
        
        detected_colors_list = sorted(list(colors_set))
        codes = []
        
        prefix = f"XHJ-{text}"
        
        # Mapping Logic
        # White -> B
        if "White" in colors_set:
            codes.append(f"{prefix} B")
        
        # Red -> H
        if "Red" in colors_set:
            codes.append(f"{prefix} H")
        
        # Green -> L
        if "Green" in colors_set:
            codes.append(f"{prefix} L")
        
        # Yellow -> U
        if "Yellow" in colors_set:
            codes.append(f"{prefix} U")
            
        # Blue -> A
        if "Blue" in colors_set:
            codes.append(f"{prefix} A")
        
        # Write to Excel
        ws.cell(row=current_row, column=1, value=text)
        ws.cell(row=current_row, column=2, value=", ".join(detected_colors_list))
        ws.cell(row=current_row, column=3, value=", ".join(codes))
        ws.cell(row=current_row, column=4, value="\n".join(img_paths)) # Newline separated paths
        
        # Center alignment
        for col in range(1, 5):
             ws.cell(row=current_row, column=col).alignment = Alignment(vertical='center', wrap_text=True)
        
        current_row += 1
            
    wb.save(excel_filename)
    print(f"Report saved to {excel_filename}")

if __name__ == "__main__":
    generate_excel_report()
