import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def check_ocr_vs_device_v7(device_excel_path: str, ocr_excel_path: str,
                           output_path: str = "Final_Check_Report_v7.xlsx"):
    """
    进行设备台账与智能识别结果的双向核对。

    规则变更：
    - 台账有现场未识别的情况，核对结果文本统一为 “该位置未识别出”（不再叠加载格式异常提示）。
    - 现场正确识别但台账存在排版格式问题时，提示文案变更为 “（注：位置格式异常）”。
    """
    try:
        # ==========================================
        # 1. 读取并清洗【设备台账数据】
        # ==========================================
        df_dev_raw = pd.read_excel(device_excel_path, header=1)
        target_cols = ["序号", "型号/规格", "站场", "位置"]

        for col in target_cols:
            if col not in df_dev_raw.columns:
                raise KeyError(f"设备表中未找到列: '{col}'，请检查表头。")

        df_dev = df_dev_raw[target_cols].dropna(subset=["序号", "位置"], how="all")
        df_dev["序号"] = pd.to_numeric(df_dev["序号"], errors="coerce")
        df_dev = df_dev.dropna(subset=["序号"])
        df_dev["序号"] = df_dev["序号"].astype(int)

        # ==========================================
        # 2. 读取并清洗【OCR 智能识别数据】
        # ==========================================
        df_ocr = pd.read_excel(ocr_excel_path)
        if "new_name" not in df_ocr.columns:
            raise KeyError(f"OCR识别表中未找到关键列: 'new_name'")

        # 提取所有识别到的有效位置集合，去除多余空白
        ocr_positions = set(df_ocr["new_name"].dropna().astype(str).str.strip().unique())

        # ==========================================
        # 3. 核心核对逻辑 (以台账为主线进行核对)
        # ==========================================
        # 将多个连续空格规范化为单个空格，确保能正常和OCR结果进行业务比对
        df_dev_clean = df_dev.copy()
        df_dev_clean["位置_干净"] = df_dev_clean["位置"].astype(str).apply(lambda x: re.sub(r'\s+', ' ', x).strip())

        pos_device_mapping = df_dev_clean.groupby("位置_干净")["型号/规格"].apply(
            lambda x: ",".join(x.astype(str))).to_dict()
        pos_count_mapping = df_dev_clean.groupby("位置_干净").size().to_dict()

        final_rows = []
        matched_ocr_positions = set()  # 记录已经被台账匹配上的识别位置

        for index, row in df_dev.iterrows():
            pos_raw = str(row["位置"])  # 带有原始排版错误（可能含连续空格/换行）

            # 内部业务对齐用的干净文本（多空格压缩为单空格，并去除两头空白）
            pos_clean = re.sub(r'\s+', ' ', pos_raw).strip()
            model = row["型号/规格"]

            # --- 异常排版检测机制 ---
            has_format_issue = False
            # 检查是否以换行符、普通空格、或不换行空格(\xa0)结尾，或者内部包含换行符/连续多空格
            if ("\n" in pos_raw or "\r" in pos_raw or
                    pos_raw.endswith(" ") or pos_raw.endswith("\xa0") or
                    bool(re.search(r'\s{2,}', pos_raw))):
                has_format_issue = True

            new_row = {
                "序号": row["序号"],
                "型号/规格": model,
                "站场": row["站场"],
                "位置": pos_raw,  # 展示原始值，方便工艺员肉眼对照修改
                "智能识别的位置": ""
            }

            # 情况 A：智能识别里【正好检测到了】这个台账位置
            if pos_clean in ocr_positions:
                new_row["智能识别的位置"] = pos_clean
                matched_ocr_positions.add(pos_clean)

                # 根据排版异常检测结果，决定是否追加新版提示语
                format_suffix = "（注：位置格式异常）" if has_format_issue else ""

                # 特判：如果台账中该位置登记了多个器材，进行多资产合并提示
                if pos_count_mapping.get(pos_clean, 1) > 1:
                    all_models = pos_device_mapping.get(pos_clean, "")
                    new_row["核对结果"] = f"该位置含有{pos_count_mapping[pos_clean]}个器材，{all_models}{format_suffix}"
                else:
                    new_row["核对结果"] = f"正确{format_suffix}"

            # 情况 B：台账有，但智能识别【没有检测出来】
            else:
                new_row["智能识别的位置"] = ""
                new_row["核对结果"] = "该位置未识别出"  # 修正：即使有排版问题，未识别出时也完全不标注异常

            final_rows.append(new_row)

        # ==========================================
        # 4. 核心核对逻辑 (找出【智能识别额外出来】的位置)
        # ==========================================
        extra_positions = ocr_positions - matched_ocr_positions

        extra_rows = []
        for extra_pos in sorted(list(extra_positions)):
            extra_rows.append({
                "序号": "",
                "型号/规格": "",
                "站场": "",
                "位置": "",
                "智能识别的位置": extra_pos,
                "核对结果": "智能识别额外出来的位置（台账无此设备）"
            })

        # ==========================================
        # 5. 组装结果并应用 openpyxl 商务美化
        # ==========================================
        df_output = pd.DataFrame(final_rows + extra_rows)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "数据防错核对"
        ws.views.sheetView[0].showGridLines = True

        # 商务颜色样式设计
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        regular_font = Font(name="Segoe UI", size=10)
        bold_font = Font(name="Segoe UI", size=10, bold=True)

        # 状态颜色高亮填充
        fill_correct = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # 淡绿（正确）
        fill_missing = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # 淡橙（未识别出）
        fill_extra = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 淡黄（额外多出）

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
        )

        ws.append(df_output.columns.tolist())

        for index, row in df_output.iterrows():
            ws.append(["" if pd.isna(v) else v for v in row.tolist()])

        # 样式与高亮渲染
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
                                      start=1):
            res_cell_value = str(ws.cell(row=row_idx, column=6).value)

            for col_idx, cell in enumerate(row, start=1):
                cell.border = thin_border

                if row_idx == 1:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.font = regular_font

                    if col_idx in [1, 3, 4, 5]:
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                    # 动态高亮背景
                    if res_cell_value.startswith("正确"):
                        if col_idx == 6: cell.fill = fill_correct
                    elif "该位置未识别出" in res_cell_value:
                        if col_idx in [4, 6]: cell.fill = fill_missing
                    elif "智能识别额外出来" in res_cell_value:
                        cell.fill = fill_extra
                        cell.font = bold_font

                    # 如果结果里包含了特殊排版隐患提示，文字颜色设为暗红加粗显示
                    if "（注：位置格式异常）" in res_cell_value and col_idx == 6:
                        cell.font = Font(name="Segoe UI", size=10, bold=True, color="C00000")

        # 设置自适应行高与列宽
        ws.row_dimensions[1].height = 28
        for r in range(2, ws.max_row + 1):
            ws.row_dimensions[r].height = 20

        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 16)

        ws.freeze_panes = "A2"
        wb.save(output_path)

        print(f"🎉 自动化双向比对完成！已应用最新精简逻辑。")
        print(f"📁 核对报告已保存至: {output_path}")

    except Exception as e:
        print(f"❌ 对比失败，原因: {e}")


if __name__ == "__main__":
    device_file = "test/test23/设备器材明细 (社棠点灯单元).xlsx"
    ocr_file = "test/test23/XHJ_OCR_Report_Final_v3.xlsx"

    check_ocr_vs_device_v7(device_file, ocr_file, "信号设备防错核对报告_v7.xlsx")