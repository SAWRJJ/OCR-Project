import pandas as pd
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def process_ocr_excel_final_v3(input_path: str, output_path: str = "XHJ_OCR_Report_Final_v3.xlsx"):
    """
    读取原始OCR Excel文件，按颜色、多心数量拆分、结构及场景过滤逻辑拆分成多行。

    规则设定：
    - template_match_res 为 1 时，生成 "小B" 代号。
    - template_match_res 为 3 时，保持不变，依然生成 "左B", "中B", "右B"。
    """
    # 1. 读取原始 Excel 文件
    df = pd.read_excel(input_path)

    rows_list = []

    # 正则表达式：用于匹配仅由 'D' 和数字组成的 matched_key (如 D1, D9)
    d_series_pattern = re.compile(r'^D\d+$', re.IGNORECASE)

    # 2. 遍历每一行，应用复合判定与颜色清除机制
    for index, row in df.iterrows():
        active_codes = []

        # 获取当前的 matched_key 并转为字符串
        m_key = str(row.get('matched_key', '')).strip()

        # 判断是否为 D+数字 的系列
        is_d_series = bool(d_series_pattern.match(m_key))

        # --- 颜色过筛与判定机制 ---

        # 1. 白色 (white 对应 B) -> 两者都考虑
        white_bg = str(row.get('color_white', '')).strip().upper()
        white_cnt = row.get('white_centers', 0)
        if white_bg in ['YES', 'TRUE'] or (pd.notna(white_cnt) and white_cnt > 0):
            active_codes.append('B')

        # 2. 红色 (red 对应 H) -> 两者都考虑
        if pd.notna(row.get('red_centers')) and row.get('red_centers', 0) > 0:
            active_codes.append('H')

        # 3. 绿色 (green 对应 L) -> 仅非D系列考虑
        if not is_d_series:
            if pd.notna(row.get('green_centers')) and row.get('green_centers', 0) > 0:
                active_codes.append('L')

        # 4. 黄色 (yellow 对应 U) -> 仅非D系列考虑 (包含特判数量为2的情况)
        if not is_d_series:
            yellow_cnt = row.get('yellow_centers', 0)
            if pd.notna(yellow_cnt):
                if yellow_cnt == 2:
                    active_codes.extend(['1U', '2U'])
                elif yellow_cnt > 0:
                    active_codes.append('U')

        # 5. 蓝色 (blue 对应 A) -> 仅D系列考虑
        if is_d_series:
            if pd.notna(row.get('blue_centers')) and row.get('blue_centers', 0) > 0:
                active_codes.append('A')

        # --- 结构与模板匹配结果判定机制 ---
        # 1. 双色控制
        if row.get('is_double') == 1:
            if 'B' not in active_codes:
                active_codes.append('B')
            active_codes.append('DB')

        # 2. 三色控制
        if row.get('is_tri') == 1:
            active_codes.extend(['AB', 'BB', 'CB'])

        # 3. 模板匹配结果控制 (template_match_res)
        tm_res = row.get('template_match_res')
        if tm_res == 1:
            active_codes.append('小B')
        elif tm_res == 2:
            active_codes.extend(['左B', '右B'])
        elif tm_res == 3:
            active_codes.extend(['左B', '中B', '右B'])  # 保持原样，不用修改
        elif tm_res == 4:
            active_codes.extend(['AB', 'BB', 'CB', 'DB'])

        # --- 行拆分与重复信息净化 ---
        if not active_codes:
            new_row = row.copy()
            new_row['color_letter'] = ''
            new_row['new_name'] = f"XHJ-{m_key}"
            rows_list.append(new_row)
        else:
            # 相同的 matched_key 拆分时，只有第一行保留完整原始数据，其余全部留空
            for i, code in enumerate(active_codes):
                if i == 0:
                    new_row = row.copy()
                    new_row['color_letter'] = code
                    new_row['new_name'] = f"XHJ-{m_key} {code}"
                else:
                    new_row = pd.Series(index=row.index, dtype=object)
                    new_row['matched_key'] = row['matched_key']
                    new_row['color_letter'] = code
                    new_row['new_name'] = f"XHJ-{m_key} {code}"

                rows_list.append(new_row)

    # 3. 重新组装 DataFrame
    df_processed = pd.DataFrame(rows_list)

    # 调整输出列顺序
    front_cols = ['new_name', 'matched_key', 'color_letter']
    other_cols = [c for c in df.columns if c not in front_cols]
    df_final = df_processed[front_cols + other_cols]

    # 4. 使用 openpyxl 进行商务美化导出
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "清洗规整表"
    ws.views.sheetView[0].showGridLines = True

    # 样式配置
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    regular_font = Font(name="Segoe UI", size=10)
    bold_font = Font(name="Segoe UI", size=10, bold=True)
    zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
    )

    ws.append(df_final.columns.tolist())

    for index, row in df_final.iterrows():
        row_values = []
        for val in row.tolist():
            if pd.isna(val):
                row_values.append("")
            elif isinstance(val, float):
                row_values.append(round(val, 4))
            else:
                row_values.append(val)
        ws.append(row_values)

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column),
                                  start=1):
        for col_idx, cell in enumerate(row, start=1):
            cell.border = thin_border
            if row_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.font = regular_font
                if col_idx == 1:
                    cell.font = bold_font
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill

                if isinstance(cell.value, (int, float)) and cell.value != "":
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                elif col_idx in [1, 2, 3]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.row_dimensions[1].height = 28
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    ws.freeze_panes = "A2"
    wb.save(output_path)
    print(f"数据已按照最终确立的命名规则处理完成。文件保存至: {output_path}")


if __name__ == "__main__":
    process_ocr_excel_final_v3("t4_ocr_report.xlsx", "XHJ_OCR_Report_Final_v3.xlsx")